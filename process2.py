import openpyxl


from openpyxl import load_workbook

wb = load_workbook('legal_cot.xlsx')
ws = wb.active
#print(ws)
print(f'The title of the Worksheet is: {ws.title}')
#print(f'The value of A2 is {ws["A2"].value}')
print(f'The value of A3 is {ws["A3"].value}')

#cell = ws['B3']
#print(f'The variable "cell" is {cell.value}')
